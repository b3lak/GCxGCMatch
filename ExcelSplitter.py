import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QCheckBox, QScrollArea, QLabel
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from PyQt5.QtWidgets import QLineEdit, QListWidget
from PyQt5.QtWidgets import QListWidgetItem, QHBoxLayout
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QCheckBox, QLabel, QListWidget, QListWidgetItem, QHBoxLayout
from PyQt5.QtCore import Qt



class ExcelSplitter(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.selectedSheetNames = set()


    def initUI(self):
        layout = QVBoxLayout()

        self.browseButton = QPushButton('Browse Excel Workbook')
        self.browseButton.clicked.connect(self.loadExcelFile)
        layout.addWidget(self.browseButton)

        self.clearAllButton = QPushButton('Clear All Selections')
        self.clearAllButton.clicked.connect(self.clearAllSelections)
        layout.addWidget(self.clearAllButton)

        self.statusLabel = QLabel('')
        layout.addWidget(self.statusLabel)

        self.counterLabel = QLabel("0/0")
        layout.addWidget(self.counterLabel)

        # Add a QLineEdit for the search functionality
        self.searchBar = QLineEdit(self)
        self.searchBar.setPlaceholderText('Search sheets...')
        self.searchBar.textChanged.connect(self.displaySheetsBasedOnSearch)
        layout.addWidget(self.searchBar)

        # Horizontal layout for two QListWidgets
        hLayout = QHBoxLayout()

        # List for all sheets
        self.sheetsListWidget = QListWidget(self)
        self.sheetsListWidget.itemChanged.connect(self.onSheetSelectionChanged)
        hLayout.addWidget(self.sheetsListWidget)

        # List for selected sheets
        self.selectedSheetsListWidget = QListWidget(self)
        hLayout.addWidget(self.selectedSheetsListWidget)
        layout.addLayout(hLayout)

        self.okButton = QPushButton('Create New Workbook')
        self.okButton.clicked.connect(self.createWorkbook)
        layout.addWidget(self.okButton)
        
        self.checkboxes = []


        self.setLayout(layout)
        self.setWindowTitle('Excel Workbook Splitter')
        self.show()

    def loadExcelFile(self):
        options = QFileDialog.Options()
        self.filePath, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        if self.filePath:
            self.excelFile = load_workbook(self.filePath)
            self.displaySheets()

    def displaySheetsBasedOnSearch(self):
        """Filter and display sheets based on search bar content."""
        filter_text = self.searchBar.text()
        self.displaySheets(filter_text)

    def displaySheets(self, filter_text=""):
        self.sheetsListWidget.clear()

        for sheet in self.excelFile.sheetnames:
            if filter_text in sheet:
                item = QListWidgetItem(sheet)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)  # Making the item checkable
                if sheet in self.selectedSheetNames:
                    item.setCheckState(Qt.Checked)
                else:
                    item.setCheckState(Qt.Unchecked)
                self.sheetsListWidget.addItem(item)
        
        
    def onSheetSelectionChanged(self, item):
        sheet_name = item.text()
        if item.checkState() == Qt.Checked:
            self.selectedSheetNames.add(sheet_name)
            self.selectedSheetsListWidget.addItem(sheet_name)
        else:
            self.selectedSheetNames.remove(sheet_name)
            items_to_remove = self.selectedSheetsListWidget.findItems(sheet_name, Qt.MatchExactly)
            if items_to_remove:
                item_to_remove = items_to_remove[0]
                row = self.selectedSheetsListWidget.row(item_to_remove)
                self.selectedSheetsListWidget.takeItem(row)
        self.updateCounter()

    def updateCounter(self):
        """Update the counter label based on selected sheets."""
        selected_count = self.selectedSheetsListWidget.count()
        total_count = self.sheetsListWidget.count()
        self.counterLabel.setText(f"{selected_count}/{total_count}")

    def filterSheets(self):
        """Filter sheet display based on search bar content."""
        filter_text = self.searchBar.text().lower()
        for checkbox in self.checkboxes:
            sheet_name = checkbox.text().lower()
            if filter_text in sheet_name:
                checkbox.setVisible(True)
            else:
                checkbox.setVisible(False)

    def clearAllSelections(self):
        """Uncheck all list items."""
        for index in range(self.sheetsListWidget.count()):
            item = self.sheetsListWidget.item(index)
            item.setCheckState(Qt.Unchecked)

    def createWorkbook(self):
        options = QFileDialog.Options()
        saveFilePath, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)", options=options)
        
        if saveFilePath:
            with pd.ExcelWriter(saveFilePath, engine='openpyxl') as writer:
                # Loop through the items in the selectedSheetsListWidget to append the sheets
                for index in range(self.selectedSheetsListWidget.count()):
                    sheet_name = self.selectedSheetsListWidget.item(index).text()
                    df = pd.read_excel(self.filePath, sheet_name=sheet_name, engine='openpyxl')
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

            book = load_workbook(saveFilePath)
            self.statusLabel.setText(f"Excel split successful! Results saved to {saveFilePath}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = ExcelSplitter()
    sys.exit(app.exec_())
